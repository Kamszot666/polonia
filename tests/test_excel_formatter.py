"""Testy czyszczenia i formatowania arkusza według standardu obróbki arkuszy Excel
przekazanego przez użytkownika."""

from openpyxl import Workbook, load_workbook

from app.export.excel_formatter import format_workbook
from config import settings

_HEADER = [
    "Lp.", "Nazwa", "Numer telefonu", "Adres e-mail", "Strona WWW",
    "Profil w mediach społecznościowych", "Data pozyskania informacji", "Status",
]


def _workbook_with(rows: list[list], tmp_path, name: str = "wejscie.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Organizacje"
    sheet.append(_HEADER)
    for row in rows:
        sheet.append(row)
    path = tmp_path / name
    workbook.save(path)
    return path


def _format(rows: list[list], tmp_path):
    source = _workbook_with(rows, tmp_path)
    output = tmp_path / "wyjscie.xlsx"
    report = format_workbook(source, output, settings)
    return load_workbook(output), report


def _row_values(sheet, row_index: int) -> dict[str, object]:
    header = [cell.value for cell in sheet[1]]
    return dict(zip(header, [cell.value for cell in sheet[row_index]]))


def test_original_file_is_not_overwritten(tmp_path):
    source = _workbook_with([[1, "Fundacja", "22 628 55 57", "biuro@a.pl", "", "", "", "done"]], tmp_path)
    before = source.read_bytes()

    format_workbook(source, tmp_path / "wyjscie.xlsx", settings)

    assert source.read_bytes() == before


def test_removes_empty_and_duplicate_rows(tmp_path):
    # Pusty wiersz musi być w środku - openpyxl i tak obcina puste wiersze na końcu arkusza.
    record = [1, "Fundacja A", "22 628 55 57", "biuro@a.pl", "", "", "18.08.2026", "done"]
    workbook, report = _format([record, [None] * 8, list(record)], tmp_path)

    assert report.empty_rows_removed == 1
    assert report.duplicate_rows_removed == 1
    assert workbook["Organizacje"].max_row == 2  # nagłówek + jeden rekord


def test_collapses_whitespace_and_trims(tmp_path):
    workbook, _ = _format([[1, "  Fundacja   Testowa  ", "22 628 55 57", "biuro@a.pl", "", "", "", "done"]], tmp_path)
    assert _row_values(workbook["Organizacje"], 2)["Nazwa"] == "Fundacja Testowa"


def test_decodes_html_entities(tmp_path):
    """Zweryfikowane realnie: nazwa z schema.org na pol.org.pl przychodzi z &quot; w środku."""
    name = "Fundacja &quot;Pomoc Polakom na Wschodzie&quot;"
    workbook, _ = _format([[1, name, "22 628 55 57", "biuro@a.pl", "", "", "", "done"]], tmp_path)
    assert _row_values(workbook["Organizacje"], 2)["Nazwa"] == 'Fundacja "Pomoc Polakom na Wschodzie"'


def test_fills_empty_cells_with_placeholder(tmp_path):
    workbook, _ = _format([[1, "Fundacja A", "22 628 55 57", "biuro@a.pl", "", "", "", "done"]], tmp_path)
    values = _row_values(workbook["Organizacje"], 2)
    assert values["Profil w mediach społecznościowych"] == settings.missing_value_placeholder


def test_treats_brak_as_empty(tmp_path):
    """Pipeline wpisuje "brak"/"nie ustalono" - wytyczne każą ujednolicić to do jednej wartości."""
    workbook, _ = _format([[1, "Fundacja A", "brak", "biuro@a.pl", "", "", "", "done"]], tmp_path)
    assert _row_values(workbook["Organizacje"], 2)["Numer telefonu"] == settings.missing_value_placeholder


def test_numbering_is_continuous_from_first_data_row(tmp_path):
    rows = [
        [7, "Fundacja A", "22 628 55 57", "a@a.pl", "", "", "", "done"],
        [None] * 8,
        [99, "Fundacja B", "22 628 55 58", "b@b.pl", "", "", "", "done"],
    ]
    workbook, _ = _format(rows, tmp_path)
    sheet = workbook["Organizacje"]
    assert [sheet.cell(row=index, column=1).value for index in (2, 3)] == [1, 2]


def test_keeps_three_most_valuable_emails_in_priority_order(tmp_path):
    cell = "jan.kowalski@a.pl, noreply@a.pl, info@a.pl, biuro@a.pl, sekretariat@a.pl, zly@@adres"
    workbook, report = _format([[1, "Fundacja A", "22 628 55 57", cell, "", "", "", "done"]], tmp_path)

    assert _row_values(workbook["Organizacje"], 2)["Adres e-mail"] == (
        "sekretariat@a.pl, biuro@a.pl, info@a.pl"
    )
    assert report.emails_removed > 0


def test_phone_labels_and_prefixes_are_removed(tmp_path):
    cell = "tel. +48 22 628 55 57, kom. 0048 514 777 541, fax 22 628 55 57"
    workbook, _ = _format([[1, "Fundacja A", cell, "a@a.pl", "", "", "", "done"]], tmp_path)

    # Stacjonarny w formacie [XX] XXX XX XX, komórkowy XXX XXX XXX, duplikat usunięty.
    assert _row_values(workbook["Organizacje"], 2)["Numer telefonu"] == "[22] 628 55 57, 514 777 541"


def test_keeps_at_most_three_phones(tmp_path):
    cell = "22 628 55 57, 22 628 55 58, 22 628 55 59, 22 628 55 60"
    workbook, report = _format([[1, "Fundacja A", cell, "a@a.pl", "", "", "", "done"]], tmp_path)

    numbers = _row_values(workbook["Organizacje"], 2)["Numer telefonu"].split(", ")
    assert len(numbers) == 3
    assert report.phones_removed == 1


def test_website_keeps_contact_subpage_without_tracking(tmp_path):
    cell = "https://a.pl/?utm_source=fb, https://a.pl/kontakt/"
    workbook, _ = _format([[1, "Fundacja A", "22 628 55 57", "a@a.pl", cell, "", "", "done"]], tmp_path)
    assert _row_values(workbook["Organizacje"], 2)["Strona WWW"] == "https://a.pl/kontakt"


def test_social_media_keeps_only_allowed_services(tmp_path):
    cell = "https://facebook.com/a/, https://vk.com/a, https://facebook.com/a"
    workbook, report = _format([[1, "Fundacja A", "22 628 55 57", "a@a.pl", "", cell, "", "done"]], tmp_path)

    assert _row_values(workbook["Organizacje"], 2)["Profil w mediach społecznościowych"] == (
        "https://facebook.com/a"
    )
    assert report.social_profiles_removed == 2


def test_removes_record_with_only_a_name(tmp_path):
    """Status i data są wpisywane zawsze, więc nie liczą się jako dane merytoryczne."""
    rows = [
        [1, "Podmiot Bez Danych", "brak", "brak", "brak", "brak", "18.08.2026", "failed"],
        [2, "Fundacja A", "22 628 55 57", "a@a.pl", "", "", "18.08.2026", "done"],
    ]
    workbook, report = _format(rows, tmp_path)

    assert report.name_only_rows_removed == 1
    assert _row_values(workbook["Organizacje"], 2)["Nazwa"] == "Fundacja A"


def test_removes_record_without_name(tmp_path):
    rows = [[1, "", "22 628 55 57", "a@a.pl", "", "", "18.08.2026", "done"]]
    _, report = _format(rows, tmp_path)
    assert report.nameless_rows_removed == 1


def test_applies_required_visual_formatting(tmp_path):
    workbook, _ = _format([[1, "Fundacja A", "22 628 55 57", "a@a.pl", "", "", "", "done"]], tmp_path)
    sheet = workbook["Organizacje"]

    assert sheet["A1"].font.bold
    assert sheet["A1"].alignment.horizontal == "center"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:H2"
    assert sheet["B2"].alignment.wrap_text
    assert sheet["B2"].alignment.horizontal == "left"
    assert sheet["B2"].alignment.vertical == "top"
    assert sheet["B2"].border.left.style == "thin"
    assert sheet.column_dimensions["B"].width > 8


def test_writes_change_report_sheet(tmp_path):
    workbook, _ = _format([[1, "Fundacja A", "22 628 55 57", "a@a.pl", "", "", "", "done"]], tmp_path)
    assert "Raport zmian" in workbook.sheetnames
    labels = [row[0] for row in workbook["Raport zmian"].iter_rows(min_row=2, values_only=True)]
    assert "Wierszy na wyjściu" in labels
