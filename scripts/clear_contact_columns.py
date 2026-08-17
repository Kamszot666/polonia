"""Czyści kolumny osoby kontaktowej (Osoba kontaktowa, telefon, e-mail) w pliku wejściowym,
żeby pipeline mógł je uzupełnić od nowa (np. po włączeniu prawdziwego NER).

Użycie: python clear_contact_columns.py <plik_wejsciowy.xlsx> <plik_wyjsciowy.xlsx>
"""

from __future__ import annotations

import sys

import openpyxl

_COLUMNS_TO_CLEAR = (
    "Osoba kontaktowa",
    "Numer telefonu do osoby kontaktowej",
    "Adres e-mail do osoby kontaktowej",
)


def main() -> None:
    if len(sys.argv) != 3:
        print("Użycie: python clear_contact_columns.py <plik_wejsciowy.xlsx> <plik_wyjsciowy.xlsx>")
        raise SystemExit(1)

    input_path, output_path = sys.argv[1], sys.argv[2]
    workbook = openpyxl.load_workbook(input_path)
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    column_indexes = [header.index(name) + 1 for name in _COLUMNS_TO_CLEAR if name in header]

    if len(column_indexes) != len(_COLUMNS_TO_CLEAR):
        missing = set(_COLUMNS_TO_CLEAR) - set(header)
        print(f"UWAGA: nie znaleziono kolumn: {missing}")

    for row in range(2, sheet.max_row + 1):
        for column_index in column_indexes:
            sheet.cell(row=row, column=column_index, value="brak")

    workbook.save(output_path)
    print(f"Wyczyszczono {len(column_indexes)} kolumn w {sheet.max_row - 1} wierszach.")
    print(f"Zapisano: {output_path}")


if __name__ == "__main__":
    main()
