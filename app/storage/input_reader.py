"""Odczyt danych wejściowych: prosta lista nazw (CSV/Excel, krok 1 algorytmu) albo już
częściowo wypełniona baza w pełnym, 19-kolumnowym schemacie (do uzupełnienia braków)."""

from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl

from app.logging.logger import logger
from app.models.schemas import ContactPerson, FieldValue, Organization, OrganizationStatus, SourceType

_NAME_COLUMN_CANDIDATES = {"nazwa", "name", "nazwa organizacji", "nazwa podmiotu"}

_MISSING_MARKERS = {"", "brak", "nie ustalono"}

_FULL_SCHEMA_HEADERS = (
    "Lp.", "Kategoria", "Branża / Typ", "Nazwa", "Adres korespondencyjny", "Województwo",
    "Numer telefonu", "Adres e-mail", "Strona WWW", "Profil w mediach społecznościowych",
    "Osoba kontaktowa", "Numer telefonu do osoby kontaktowej", "Adres e-mail do osoby kontaktowej",
    "Data pozyskania informacji", "Krótka charakterystyka podmiotu", "URL źródła", "KRS", "REGON", "NIP",
)

# Konwencja potwierdzona w realnym pliku użytkownika: "Jan Kowalski (Prezes)".
_PERSON_WITH_POSITION_PATTERN = re.compile(r"^(?P<name>.*?)\s*\((?P<position>[^)]+)\)\s*$")


def read_organization_names(input_path: Path) -> list[str]:
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_from_excel(input_path)
    if suffix == ".csv":
        return _read_from_csv(input_path)
    raise ValueError(f"Nieobsługiwany format pliku wejściowego: {suffix}")


def _read_from_excel(path: Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    name_column_index = _find_name_column_index(header)

    names = [
        str(row[name_column_index]).strip()
        for row in rows
        if name_column_index < len(row) and row[name_column_index]
    ]
    workbook.close()
    return _deduplicate(names)


def _read_from_csv(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header = next(reader, None)
        name_column_index = _find_name_column_index(header)
        names = [
            row[name_column_index].strip()
            for row in reader
            if len(row) > name_column_index and row[name_column_index].strip()
        ]
    return _deduplicate(names)


def _find_name_column_index(header: tuple | list | None) -> int:
    if header:
        for index, cell in enumerate(header):
            if cell and str(cell).strip().lower() in _NAME_COLUMN_CANDIDATES:
                return index
    logger.warning("Nie znaleziono nagłówka z nazwą organizacji - używam pierwszej kolumny")
    return 0


def _deduplicate(names: list[str]) -> list[str]:
    seen: dict[str, None] = {}
    for name in names:
        seen.setdefault(name, None)
    return list(seen)


def is_full_schema_workbook(path: Path) -> bool:
    """Rozpoznaje plik zgodny z pełnym, 19-kolumnowym schematem bazy (do uzupełnienia braków),
    zamiast zwykłej listy nazw - sprawdzając rzeczywisty nagłówek, a nie samo rozszerzenie pliku."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        header = next(workbook.active.iter_rows(values_only=True), None)
    finally:
        workbook.close()
    if not header:
        return False
    normalized = tuple(str(cell).strip() if cell is not None else "" for cell in header)
    return normalized[: len(_FULL_SCHEMA_HEADERS)] == _FULL_SCHEMA_HEADERS


def read_organizations_from_workbook(path: Path) -> list[Organization]:
    """Wczytuje istniejącą, częściowo wypełnioną bazę (pełny schemat 19 kolumn). Dane już znane
    trafiają do modelu jako wysokopewne FieldValue (SourceType.MANUAL_INPUT, confidence=1.0) i
    nie zostaną nadpisane przez pipeline; komórki "brak"/"nie ustalono"/puste stają się brakami
    do uzupełnienia."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    next(rows, None)  # nagłówek

    organizations = [_row_to_organization(row) for row in rows if any(row)]
    workbook.close()
    return organizations


def _row_to_organization(row: tuple) -> Organization:
    padded = row + (None,) * max(0, len(_FULL_SCHEMA_HEADERS) - len(row))
    (
        _lp, kategoria, branza, nazwa, adres, wojewodztwo, telefon, email, www, social,
        osoba, telefon_osoby, email_osoby, data_pozyskania, opis, url_zrodla, krs, regon, nip,
    ) = padded[: len(_FULL_SCHEMA_HEADERS)]

    person_name, person_position = _split_person(_clean(osoba))

    return Organization(
        input_name=_clean(nazwa) or "",
        name=_field(_clean(nazwa)),
        address=_field(_clean(adres)),
        voivodeship=_field(_clean(wojewodztwo)),
        phone=_field(_clean(telefon)),
        email=_field(_clean(email)),
        website=_field(_clean(www)),
        social_media=_field(_clean(social)),
        contact_person=ContactPerson(
            name=_field(person_name),
            position=_field(person_position),
            phone=_field(_clean(telefon_osoby)),
            email=_field(_clean(email_osoby)),
        ),
        description=_field(_clean(opis)),
        krs=_field(_clean(krs)),
        regon=_field(_clean(regon)),
        nip=_field(_clean(nip)),
        category=_clean(kategoria),
        industry=_field(_clean(branza)),
        origin_source_url=_clean(url_zrodla),
        date_acquired=_clean(data_pozyskania),
        status=OrganizationStatus.PENDING,
    )


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in _MISSING_MARKERS:
        return None
    return text


def _field(value: str | None) -> FieldValue:
    if value is None:
        return FieldValue()
    return FieldValue(value=value, source_type=SourceType.MANUAL_INPUT, evidence="plik wejściowy", confidence=1.0)


def _split_person(text: str | None) -> tuple[str | None, str | None]:
    if text is None:
        return None, None
    match = _PERSON_WITH_POSITION_PATTERN.match(text)
    if match:
        return match.group("name").strip() or None, match.group("position").strip() or None
    return text, None
