"""Zapis wyników do Excela przez openpyxl (pkt 11 algorytmu).

Układ i nazwy kolumn 1-19 oraz konwencja wypełniania braków ("brak" / "nie ustalono"
zamiast pustej komórki) odpowiadają dostarczonej „Procedurze tworzenia bazy danych
w Excelu” oraz realnemu plikowi bazy przekazanemu przez użytkownika (Lp. ... KRS,
REGON, NIP). Kolumny 20-21 (Status, Źródła i pewność) to rozszerzenie wymagane przez
instrukcję Polonia (pkt 9: każde pole musi mieć przypisane źródło i confidence) -
procedura ogólna ich nie przewiduje, ale nie jest z nimi sprzeczna.

"Kategoria" i "Branża / Typ" nie są w pełni automatycznie wykrywane - Branża bywa
wypełniana z przeważającego przedmiotu działalności (PKD) z API KRS, gdy dostępne;
w pozostałych przypadkach zostaje "brak" do ręcznego uzupełnienia.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.export.excel_formatter import format_workbook
from app.logging.logger import logger
from app.models.schemas import FieldValue, Organization
from config import Settings, settings

_BRAK = "brak"
_NIE_USTALONO = "nie ustalono"

_HEADERS = (
    "Lp.", "Kategoria", "Branża / Typ", "Nazwa", "Adres korespondencyjny", "Województwo",
    "Numer telefonu", "Adres e-mail", "Strona WWW", "Profil w mediach społecznościowych",
    "Osoba kontaktowa", "Numer telefonu do osoby kontaktowej", "Adres e-mail do osoby kontaktowej",
    "Data pozyskania informacji", "Krótka charakterystyka podmiotu", "URL źródła", "KRS", "REGON", "NIP",
    "Status", "Źródła i pewność",
)


def export_to_excel(organizations: list[Organization], settings: Settings = settings) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Organizacje"
    sheet.append(_HEADERS)

    for index, org in enumerate(organizations, start=1):
        sheet.append(_organization_to_row(index, org))

    _autosize_columns(sheet)

    settings.output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(settings.output_xlsx_path)
    logger.info(f"Zapisano {len(organizations)} rekordów do {settings.output_xlsx_path}")

    if not settings.format_output_workbook:
        return settings.output_xlsx_path

    formatted_path = settings.output_xlsx_path.with_name(
        f"{settings.output_xlsx_path.stem}{settings.formatted_output_suffix}"
        f"{settings.output_xlsx_path.suffix}"
    )
    report = format_workbook(settings.output_xlsx_path, formatted_path, settings)
    for label, value in report.as_lines():
        logger.info(f"  {label}: {value}")
    return formatted_path


def _organization_to_row(lp: int, org: Organization) -> list[str]:
    person = org.contact_person
    contact_person_display = _format_contact_person(org)
    return [
        lp,
        org.category or _BRAK,
        _value_or(org.industry, _BRAK),
        # Nazwa z KRS/schema.org bywa nieustalona, ale nazwa wejściowa jest zawsze - bez tego
        # fallbacku rekord trafiał do arkusza bez nazwy i przepadał przy czyszczeniu, mimo że
        # miał komplet danych kontaktowych.
        _value_or(org.name, org.input_name or _BRAK),
        _value_or(org.address, _BRAK),
        _value_or(org.voivodeship, _BRAK),
        _value_or(org.phone, _BRAK),
        _value_or(org.email, _BRAK),
        _value_or(org.website, _BRAK),
        _value_or(org.social_media, _BRAK),
        contact_person_display or _NIE_USTALONO,
        _value_or(person.phone, _BRAK),
        _value_or(person.email, _BRAK),
        org.date_acquired or _BRAK,
        _value_or(org.description, _BRAK),
        org.origin_source_url or _BRAK,
        _value_or(org.krs, _BRAK),
        _value_or(org.regon, _BRAK),
        _value_or(org.nip, _BRAK),
        org.status.value,
        _summarize_sources(org),
    ]


def _format_contact_person(org: Organization) -> str:
    name = org.contact_person.name.value
    position = org.contact_person.position.value
    if name and position:
        return f"{name} ({position})"
    return name or ""


def _value_or(field: FieldValue, fallback: str) -> str:
    return field.value if not field.is_empty else fallback


def _summarize_sources(org: Organization) -> str:
    labeled_fields: list[tuple[str, FieldValue]] = [
        ("nazwa", org.name), ("adres", org.address), ("województwo", org.voivodeship),
        ("telefon", org.phone), ("e-mail", org.email), ("WWW", org.website),
        ("media społ.", org.social_media),
        ("osoba", org.contact_person.name), ("stanowisko", org.contact_person.position),
        ("telefon osoby", org.contact_person.phone), ("e-mail osoby", org.contact_person.email),
        ("opis", org.description), ("branża", org.industry),
        ("KRS", org.krs), ("REGON", org.regon), ("NIP", org.nip),
    ]
    parts = []
    for label, field in labeled_fields:
        if field.is_empty:
            continue
        source = field.source_type.value if field.source_type else "?"
        url = field.source_url or "?"
        parts.append(f"{label}: {source} @ {url} (conf={field.confidence:.2f})")
    return "; ".join(parts)


def _autosize_columns(sheet: Worksheet, max_width: int = 60) -> None:
    for index in range(1, len(_HEADERS) + 1):
        column_letter = get_column_letter(index)
        max_length = max(
            (len(str(cell.value)) for cell in sheet[column_letter] if cell.value is not None),
            default=10,
        )
        sheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)
