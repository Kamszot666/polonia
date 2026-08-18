"""Czyszczenie i formatowanie arkusza wynikowego według standardu obróbki arkuszy.

Realizuje wytyczne przekazane przez użytkownika („formatowanie i obróbka arkuszy Excel"):
usuwanie pustych i zduplikowanych wierszy, ujednolicenie telefonów i adresów e-mail,
ograniczenie ich liczby do trzech najbardziej wartościowych, oczyszczenie adresów WWW
i profili społecznościowych, ciągła numeracja oraz wygląd arkusza (nagłówki, filtr,
zamrożony wiersz, obramowania, naprzemienne tło, dopasowanie szerokości).

Wynik zapisywany jest zawsze do nowego pliku - oryginał zostaje nietknięty, zgodnie
z wytycznymi.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from html import unescape
from pathlib import Path

import phonenumbers
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.extract.email_extractor import rank_emails
from app.extract.phone_extractor import is_callable_number
from app.extract.social_media_extractor import clean_social_url, platform_of
from app.logging.logger import logger
from config import Settings, settings

_MULTIPLE_SPACES = re.compile(r"\s+")
# Etykiety doklejane do numerów na stronach - wytyczne każą je usunąć przed formatowaniem.
_PHONE_LABELS = re.compile(
    r"\b(tel\.?|telefon kontaktowy|telefon|kom\.?|fax\.?|faks\.?|gsm|mobile|mob\.?)\b[:\s]*",
    re.IGNORECASE,
)
_TRACKING_QUERY = re.compile(r"[?&](utm_[^=&]*|fbclid|gclid|mc_cid|mc_eid|_ga)=[^&]*", re.IGNORECASE)

# Wartości, które w praktyce znaczą "pusto" - wchodzą pod placeholder z wytycznych.
_EMPTY_MARKERS = {"", "-", "--", "brak", "brak danych", "nie ustalono", "nie znaleziono", "n/d", "nd"}

_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
_BANDED_FILL = PatternFill("solid", fgColor="F2F2F2")
_THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


@dataclass(slots=True)
class CleanupReport:
    """Raport z wykonanych zmian - wytyczne wymagają go po każdym czyszczeniu arkusza."""

    rows_in: int = 0
    rows_out: int = 0
    empty_rows_removed: int = 0
    duplicate_rows_removed: int = 0
    nameless_rows_removed: int = 0
    name_only_rows_removed: int = 0
    emails_removed: int = 0
    phones_removed: int = 0
    urls_cleaned: int = 0
    social_profiles_removed: int = 0
    cells_filled: int = 0
    notes: list[str] = field(default_factory=list)

    def as_lines(self) -> list[tuple[str, int | str]]:
        return [
            ("Wierszy na wejściu", self.rows_in),
            ("Wierszy na wyjściu", self.rows_out),
            ("Usunięto pustych wierszy", self.empty_rows_removed),
            ("Usunięto duplikatów", self.duplicate_rows_removed),
            ("Usunięto rekordów bez nazwy", self.nameless_rows_removed),
            ("Usunięto rekordów z samą nazwą", self.name_only_rows_removed),
            ("Usunięto nadmiarowych/błędnych adresów e-mail", self.emails_removed),
            ("Usunięto nadmiarowych/błędnych numerów telefonu", self.phones_removed),
            ("Oczyszczono adresów WWW", self.urls_cleaned),
            ("Usunięto profili spoza dozwolonych serwisów", self.social_profiles_removed),
            ("Komórek uzupełnionych placeholderem", self.cells_filled),
        ]


def format_workbook(
    input_path: Path, output_path: Path, custom_settings: Settings = settings
) -> CleanupReport:
    """Czyta arkusz, czyści dane według wytycznych i zapisuje wynik do nowego pliku."""
    source = load_workbook(input_path)
    sheet = _pick_data_sheet(source)
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    if not rows:
        raise ValueError(f"Arkusz {input_path} jest pusty - nie ma czego formatować")

    header = [_clean_text(value) for value in rows[0]]
    report = CleanupReport(rows_in=len(rows) - 1)
    records = _clean_records(rows[1:], header, report, custom_settings)
    report.rows_out = len(records)

    target = Workbook()
    target_sheet = target.active
    target_sheet.title = sheet.title
    target_sheet.append(header)
    for number, record in enumerate(records, start=1):
        target_sheet.append(_with_row_number(record, header, number))

    _apply_visual_format(target_sheet, len(header), len(records))
    _append_report_sheet(target, report)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    target.save(output_path)
    logger.info(f"Sformatowano arkusz: {report.rows_in} -> {report.rows_out} rekordów, zapis do {output_path}")
    return report


def _pick_data_sheet(workbook: Workbook) -> Worksheet:
    """Wytyczne: zachowaj tylko jeden arkusz roboczy, jeśli pozostałe są puste."""
    non_empty = [sheet for sheet in workbook.worksheets if sheet.max_row > 1]
    return non_empty[0] if non_empty else workbook.worksheets[0]


def _clean_records(
    raw_rows: list[list], header: list[str], report: CleanupReport, config: Settings
) -> list[list[str]]:
    kinds = [_column_kind(name) for name in header]
    name_index = kinds.index("name") if "name" in kinds else None
    bookkeeping = _bookkeeping_indexes(header)
    seen: set[tuple[str, ...]] = set()
    records: list[list[str]] = []

    for raw_row in raw_rows:
        row = [_clean_text(value) for value in _padded(raw_row, len(header))]
        row = [_normalize_marker(value) for value in row]

        if not any(row):
            report.empty_rows_removed += 1
            continue

        for index, kind in enumerate(kinds):
            row[index] = _clean_by_kind(row[index], kind, report, config)

        if name_index is not None:
            if not row[name_index]:
                report.nameless_rows_removed += 1
                report.notes.append(f"rekord bez nazwy pominięty (wiersz: {row[:4]})")
                continue
            if _is_technical_record(row[name_index]):
                report.nameless_rows_removed += 1
                report.notes.append(f"rekord techniczny pominięty: {row[name_index]!r}")
                continue
            if config.drop_records_without_data and not _has_data_beyond_name(
                row, kinds, name_index, bookkeeping
            ):
                report.name_only_rows_removed += 1
                report.notes.append(f"rekord bez danych poza nazwą: {row[name_index]}")
                continue

        # Duplikat liczony bez numeru porządkowego - ten i tak jest nadawany od nowa.
        fingerprint = tuple(value for index, value in enumerate(row) if kinds[index] != "number")
        if fingerprint in seen:
            report.duplicate_rows_removed += 1
            continue
        seen.add(fingerprint)

        for index, value in enumerate(row):
            if not value and kinds[index] != "number":
                row[index] = config.missing_value_placeholder
                report.cells_filled += 1
        records.append(row)

    return records


def _clean_by_kind(value: str, kind: str, report: CleanupReport, config: Settings) -> str:
    if not value:
        return value
    if kind == "email":
        return _clean_email_cell(value, report, config)
    if kind == "phone":
        return _clean_phone_cell(value, report, config)
    if kind == "website":
        return _clean_website_cell(value, report, config)
    if kind == "social":
        return _clean_social_cell(value, report, config)
    return value


def _clean_email_cell(value: str, report: CleanupReport, config: Settings) -> str:
    found = list(dict.fromkeys(re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", value)))
    kept = rank_emails(found, config.max_emails_per_organization)
    report.emails_removed += max(0, len(found) - len(kept))
    return config.contact_list_separator.join(kept)


def _clean_phone_cell(value: str, report: CleanupReport, config: Settings) -> str:
    stripped = _PHONE_LABELS.sub(" ", value)
    formatted: list[str] = []
    seen: set[str] = set()
    total = 0
    for match in phonenumbers.PhoneNumberMatcher(stripped, "PL"):
        total += 1
        if not is_callable_number(match.number):
            continue
        national = str(match.number.national_number)
        if national in seen:
            continue
        seen.add(national)
        if len(formatted) < config.max_phones_per_organization:
            formatted.append(_format_polish_number(match.number))
    report.phones_removed += max(0, total - len(formatted))
    return config.contact_list_separator.join(formatted)


def _format_polish_number(number: phonenumbers.PhoneNumber) -> str:
    """Wytyczne: komórkowe jako XXX XXX XXX, stacjonarne jako [42] XXX XX XX, bez prefiksu +48."""
    national = str(number.national_number)
    if len(national) != 9:
        return national
    if phonenumbers.number_type(number) == phonenumbers.PhoneNumberType.MOBILE:
        return f"{national[:3]} {national[3:6]} {national[6:]}"
    return f"[{national[:2]}] {national[2:5]} {national[5:7]} {national[7:]}"


def _clean_website_cell(value: str, report: CleanupReport, config: Settings) -> str:
    urls = [part for part in re.split(r"[,\s]+", value) if part.startswith(("http://", "https://", "www."))]
    if not urls:
        return value
    cleaned = [_TRACKING_QUERY.sub("", url).rstrip("/?&") for url in urls]
    unique = list(dict.fromkeys(cleaned))
    # Wytyczne: pozostaw jedną stronę, najlepiej podstronę z kontaktami.
    best = next((url for url in unique if "kontakt" in url.lower() or "contact" in url.lower()), unique[0])
    if len(unique) > 1 or best != value.strip():
        report.urls_cleaned += 1
    return best


def _clean_social_cell(value: str, report: CleanupReport, config: Settings) -> str:
    urls = [part for part in re.split(r"[,\s]+", value) if part.startswith(("http://", "https://"))]
    kept: list[str] = []
    seen_platforms: set[str] = set()
    for url in urls:
        platform = platform_of(url)
        if platform is None or platform in seen_platforms:
            continue
        seen_platforms.add(platform)
        kept.append(clean_social_url(url))
    report.social_profiles_removed += max(0, len(urls) - len(kept))
    return config.contact_list_separator.join(kept)


def _column_kind(header_name: str) -> str:
    lowered = header_name.lower()
    if lowered.startswith("lp"):
        return "number"
    if "społecznościow" in lowered or "social" in lowered:
        return "social"
    if "e-mail" in lowered or "email" in lowered or "mail" in lowered:
        return "email"
    if "telefon" in lowered or "tel." in lowered:
        return "phone"
    if "www" in lowered or "strona" in lowered:
        return "website"
    if "nazwa" in lowered:
        return "name"
    return "text"


# Kolumny, które pipeline wypełnia zawsze (status, data, spis źródeł) - bez ich wykluczenia
# żaden rekord nie zostałby uznany za "samą nazwę bez danych", bo formalnie coś w nich stoi.
_BOOKKEEPING_HEADERS = ("status", "data pozyskania", "źródła i pewność", "kategoria", "url źródła")


def _bookkeeping_indexes(header: list[str]) -> set[int]:
    return {
        index
        for index, name in enumerate(header)
        if any(marker in name.lower() for marker in _BOOKKEEPING_HEADERS)
    }


def _has_data_beyond_name(
    row: list[str], kinds: list[str], name_index: int, bookkeeping: set[int]
) -> bool:
    """Wytyczne: usuń rekordy zawierające wyłącznie nazwę bez żadnych pozostałych danych."""
    for index, value in enumerate(row):
        if index == name_index or index in bookkeeping or kinds[index] == "number" or not value:
            continue
        return True
    return False


def _with_row_number(record: list[str], header: list[str], number: int) -> list:
    """Wytyczne: pierwsza kolumna to numer porządkowy, ciągły, od drugiego wiersza."""
    row = list(record)
    if header and _column_kind(header[0]) == "number":
        row[0] = number
    return row


def _apply_visual_format(sheet: Worksheet, column_count: int, row_count: int) -> None:
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = header_alignment
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER

    for row_index in range(2, row_count + 2):
        for cell in sheet[row_index]:
            cell.alignment = body_alignment
            cell.border = _THIN_BORDER
            if row_index % 2 == 0:
                cell.fill = _BANDED_FILL
        # Wysokość automatyczna - Excel dopasowuje ją sam przy włączonym zawijaniu tekstu.
        sheet.row_dimensions[row_index].height = None

    if column_count:
        last_column = get_column_letter(column_count)
        sheet.auto_filter.ref = f"A1:{last_column}{row_count + 1}"
    sheet.freeze_panes = "A2"
    _autosize_columns(sheet, column_count)


def _autosize_columns(sheet: Worksheet, column_count: int, max_width: int = 55) -> None:
    for index in range(1, column_count + 1):
        letter = get_column_letter(index)
        longest = max(
            (len(line) for cell in sheet[letter] if cell.value is not None
             for line in str(cell.value).split("\n")),
            default=10,
        )
        sheet.column_dimensions[letter].width = min(max(longest + 2, 8), max_width)


def _append_report_sheet(workbook: Workbook, report: CleanupReport) -> None:
    sheet = workbook.create_sheet("Raport zmian")
    sheet.append(["Pozycja", "Wartość"])
    for label, value in report.as_lines():
        sheet.append([label, value])
    if report.notes:
        sheet.append([])
        sheet.append(["Szczegóły"])
        for note in report.notes:
            sheet.append([note])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    sheet.column_dimensions["A"].width = 48
    sheet.column_dimensions["B"].width = 14


def _padded(row: list, length: int) -> list:
    return list(row) + [None] * (length - len(row))


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    # Zweryfikowane realnie: nazwa z schema.org na pol.org.pl trafiała do arkusza jako
    # 'Fundacja &quot;Pomoc Polakom na Wschodzie&quot;' - encje muszą zostać rozwinięte.
    return _MULTIPLE_SPACES.sub(" ", unescape(str(value))).strip()


def _normalize_marker(value: str) -> str:
    return "" if value.lower() in _EMPTY_MARKERS else value


def _is_technical_record(name: str) -> bool:
    lowered = name.lower()
    return lowered in _EMPTY_MARKERS or lowered in {"nazwa", "test", "przykład", "example"}
